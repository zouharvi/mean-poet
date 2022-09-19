#!/usr/bin/env python3

import sys
sys.path.append("src")
import json
import argparse
import random
from openpyxl import Workbook
from openpyxl.styles import Alignment, Color, PatternFill
from xlsx_constants import *

FIELDNAMES = [
    "title",
    "t_key",
    "translator",
    "translator_level",
    "stanza_i",
    "src",
    "hyp",
]

ATTRIBUTES = [
    "Meaning", "Poeticness", "Overall"
]


def ord_to_col(i):
    if i < 0:
        raise Exception(f"Attempted to transform {i} into a column name")
    if i < 26:
        return chr(ord("A") + i)
    if i < 26 * 26:
        return ord_to_col(i // 26 - 1) + ord_to_col(i % 26)
    raise Exception(f"Too large a column number {i}")


def get_height_for_row(sheet, row_number):
    row = list(sheet.rows)[row_number - 1]
    # maximum from all cells in a row
    return max([30] + [
        # get line + long line count
        (
            cell.value.count("\n") +
            len([x for x in cell.value.split("\n") if len(x.strip()) > 60])
        ) * 19 if cell.value is not None else 0
        for cell in row
    ])


def add_sheet(workbook, poem, poem_i, t_keys, reference_first):
    sheet = workbook.create_sheet(f"{poem_i:0>2}")
    sheet["A1"].value = "Source"
    sheet["A1"].font = FONT_BOLD

    if reference_first:
        sheet["B1"].value = "Reference"
        sheet["B1"].font = FONT_BOLD
        

    for i in range(len(t_keys)):
        for a_i, a in enumerate(ATTRIBUTES):
            col = ord_to_col(i * (len(ATTRIBUTES) + 1) + a_i + 2)
            cell = sheet[f"{col}1"]
            cell.value = a
            cell.font = FONT_BOLD
            cell.alignment = Alignment(
                textRotation=90, horizontal="center"
            )
            sheet.column_dimensions[col].width = 3

        # store translation source
        sheet[f"{ord_to_col(i * (len(ATTRIBUTES) + 1) + 1)}2"].value = poem[t_keys[i]]["translator"]

    sheet.row_dimensions[1].height = 85
    sheet.freeze_panes = sheet["B3"]

    # hide translation source
    sheet.row_dimensions[2].hidden = True

    for t_key_i, t_key in enumerate(t_keys):
        for stanza_i, stanza_hyp in enumerate([poem[t_key]["title"]] + poem[t_key]["poem"].split("\n\n")):
            col = ord_to_col(t_key_i * (len(ATTRIBUTES) + 1) + 1)

            for a_i, a in enumerate(ATTRIBUTES):
                col_a = ord_to_col(t_key_i * (len(ATTRIBUTES) + 1) +  a_i + 2)
                cell_a = sheet[f"{col_a}{stanza_i+3}"]
                cell_a.font = FONT_BOLD

            cell = sheet[f"{col}{stanza_i+3}"]
            cell.value = stanza_hyp.strip()
            if stanza_i % 2 == 0:
                cell.fill = FILL_PAIRS[t_key_i][0]
            else:
                cell.fill = FILL_PAIRS[t_key_i][1]
            cell.font = FONT_NORMAL
            sheet.column_dimensions[col].width = 45
            cell.alignment = Alignment(wrap_text=True)

        # stanza_i+4 points to the last empty row
        for a_i, a in enumerate(ATTRIBUTES):
            col_a = ord_to_col(t_key_i * (len(ATTRIBUTES) + 1) + 1 + a_i)
            cell_a = sheet[f"{col_a}{stanza_i+4}"]
            cell_a.font = FONT_BOLD

    for stanza_i, stanza_hyp in enumerate([poem["title"]] + poem["poem"].split("\n\n")):
        cell = sheet[f"A{stanza_i+3}"]
        cell.value = stanza_hyp.strip()
        if stanza_i % 2 == 0:
            cell.fill = FILL_0A
        else:
            cell.fill = FILL_0B
        cell.font = FONT_NORMAL

        sheet.row_dimensions[
            stanza_i + 3
        ].height = get_height_for_row(sheet, stanza_i + 3)
        sheet.column_dimensions["A"].width = 45

    # stanza_i+3 points to the last empty row
    cell = sheet[f"A{stanza_i+4}"]
    cell.value = "Overall"
    cell.font = FONT_BOLD

    if reference_first:
        # remove rating for those columns
        sheet.column_dimensions['C'].hidden= True
        sheet.column_dimensions['D'].hidden= True
        sheet.column_dimensions['E'].hidden= True

if __name__ == "__main__":
    args = argparse.ArgumentParser()
    args.add_argument(
        "-d", "--dataset", default="computed/dataset.jsonl",
        help="Path to dataset file"
    )
    args.add_argument(
        "-dm", "--dataset-meta", default="computed/dataset_meta.jsonl",
        help="Path to dataset meta file"
    )
    args.add_argument(
        "-o", "--out", default="computed/dataset.xlsx",
        help="Path to annotation dataset xlsx file"
    )
    args.add_argument(
        "-r", "--reference", action="store_true",
        help="Mark which one is reference"
    )
    args.add_argument(
        "-s", "--seed", type=int, default=0,
    )
    args = args.parse_args()

    with open(args.dataset, "r") as f:
        data = [json.loads(x) for x in f.readlines()]

    # generate XLSX document
    workbook = Workbook()
    # remove default sheet
    workbook.remove(workbook.active)

    for poem_i, poem in enumerate(data):
        poem["num"] = poem_i

    # seed twice
    random.seed(args.seed)

    # TODO: choose 10 random poems for now
    data = random.sample(data, k=10)

    random.seed(args.seed)

    for poem in data:
        t_keys = [x for x in poem.keys() if "translation-" in x]

        if args.reference:
            t_key_ref = t_keys[0]
            t_keys = t_keys[1:]
            random.shuffle(t_keys)
            # store the specific order
            t_keys = [t_key_ref] + t_keys
            poem["order_ref"] = True
        else:
            random.shuffle(t_keys)
            # store the specific order
            poem["order_ref"] = False
            
        poem["order"] = t_keys

        add_sheet(
            workbook, poem, poem["num"],
            t_keys=t_keys, reference_first=args.reference,
        )

    workbook.save(args.out)

    with open(args.dataset_meta, "w") as f:
        f.writelines([json.dumps(x, ensure_ascii=False) + "\n" for x in data])