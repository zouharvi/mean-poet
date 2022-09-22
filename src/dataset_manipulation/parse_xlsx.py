#!/usr/bin/env python3

import sys
sys.path.append("src")
from utils import json_read, json_readl
from openpyxl import load_workbook
import json
from pathlib import Path
import numpy as np

# override
UIDs = [
    "dataset_s0_vilem", "dataset_s0_eleanor",
]
DATA = []

COL_DONE = ["B", "F", "J", "N", "R", "V", "Z"]
COL_DONE_RATING = [
    ["C", "D", "E"],
    ["G", "H", "I"],
    ["K", "L", "M"],
    ["O", "P", "Q"],
    ["S", "T", "U"],
    ["W", "X", "Y"],
    ["AA", "AB", "AC"],
]
ATTRIBUTES = [
    "meaning", "poeticness", "overall"
]

for uid in UIDs:
    if not Path(f"data_annotations/{uid}.xlsx").is_file():
        continue
    # data_only allows reading formula results (via cache)
    wb = load_workbook(f"data_annotations/{uid}.xlsx", data_only=True)

    data_meta = json_readl("data_annotations/dataset_meta.jsonl")

    for sheetname, poem_meta in zip(wb.sheetnames, data_meta):
        # invalid poem (misaligned)
        if sheetname == "10":
            continue
        print(sheetname, poem_meta["order"])
        sheet_done = wb[sheetname]
        row_start = None
        row_end = None
        for row in range(2, 50):
            is_empty = sheet_done[f"B{row}"].value is None
            if is_empty and row_start is not None:
                row_end = row
                break
            if not is_empty and row_start is None:
                # +1 for the hidden row with system names
                row_start = row + 1

        row_rating = None
        for row in range(row_end, 50):
            is_empty = sheet_done[f"A{row}"].value is None
            if is_empty:
                row_rating = row -1
                break

        print(row_start, row_rating, row_end)

        line_doc = {
            "uid": uid,
            "poem_i": sheetname,
            "rating": {},
            "lines": [],
            "lang": poem_meta["lang"],
        }

        for translation_i, translation_name in enumerate(poem_meta["order"]):
            rating = {}
            for rating_i, rating_col in enumerate(COL_DONE_RATING[translation_i]):
                attribute_name = ATTRIBUTES[rating_i]
                rating[attribute_name] = sheet_done[f"{rating_col}{row_rating}"].value

            system = poem_meta[translation_name]
            line_doc["rating"][system["translator"]] = rating
                
        # row end points to the first empty after segment
        for row in range(row_start, row_end):
            line_transl = {
            }
            for translation_i, translation_name in enumerate(poem_meta["order"]):
                line_system = {
                    "orig": str(sheet_done[f"{COL_DONE[translation_i]}{row}"].value),
                }
                rating = {}
                for rating_i, rating_col in enumerate(COL_DONE_RATING[translation_i]):
                    attribute_name = ATTRIBUTES[rating_i]
                    rating[attribute_name] = sheet_done[f"{rating_col}{row}"].value
                line_system["rating"] = rating
                line_system["translator_level"] = poem_meta[translation_name]["translator_level"]

                system = poem_meta[translation_name]
                line_transl[system["translator"]] = line_system

            line_doc["lines"].append(line_transl)
        DATA.append(line_doc)
    
# not using save_json because of cache/encoding issue?
with open("data_annotations/parsed.json", "w") as f:
    json.dump(DATA, f, indent=4, ensure_ascii=False)
