from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Color, PatternFill, Font

FILL_0A = PatternFill(
    patternType='solid',
    fill_type='solid', fgColor=Color('aaaaaa')
)
FILL_0B = PatternFill(
    patternType='solid',
    fill_type='solid', fgColor=Color('dddddd')
)
FILL_1A = PatternFill(
    patternType='solid',
    fill_type='solid', fgColor=Color('ffc488')
)
FILL_1B = PatternFill(
    patternType='solid',
    fill_type='solid', fgColor=Color('fceee0')
)
FILL_2A = PatternFill(
    patternType='solid',
    fill_type='solid', fgColor=Color('8574fb')
)
FILL_2B = PatternFill(
    patternType='solid',
    fill_type='solid', fgColor=Color('eae8fc')
)
FILL_3A = PatternFill(
    patternType='solid',
    fill_type='solid', fgColor=Color('e9ff6d')
)
FILL_3B = PatternFill(
    patternType='solid',
    fill_type='solid', fgColor=Color('fcfcdb')
)
FILL_4A = PatternFill(
    patternType='solid',
    fill_type='solid', fgColor=Color('3c974b')
)
FILL_4B = PatternFill(
    patternType='solid',
    fill_type='solid', fgColor=Color('cbf2d1')
)
FILL_BLANK = PatternFill(
    patternType="solid",
    fill_type='solid', fgColor=Color('eda1a1')
)

FILL_PAIRS = [
    (FILL_1A, FILL_1B),
    (FILL_2A, FILL_2B),
    (FILL_3A, FILL_3B),
    (FILL_4A, FILL_4B),
]*2


THIN_BORDER_ALL = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
MEDIUM_BORDER_RIGHT = Border(
    left=Side(style='thin'),
    right=Side(style='medium'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
MEDIUM_BORDER_ALL = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)
THICK_BORDER_RIGHT = Border(
    left=Side(style='thin'),
    right=Side(style='thick'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
THICK_BORDER_BOTTOM = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thick')
)

FONT_BOLD = Font(bold=True, name="Arial")
FONT_NORMAL = Font(bold=False, name="Arial")